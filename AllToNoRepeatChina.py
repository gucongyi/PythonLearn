#!/usr/bin/python
# -*- coding: utf-8 -*-
#第一行注释设置全局编码
#读取all.txt
listAllfile=[]
listAllfileSplit=[]
allfile=open("all.txt","r",encoding="utf-8")
listAllfile=list(allfile)
allfile.close()
for i in range(0,len(listAllfile)):#根据|切分
    listTemp = []
    indexFirst=listAllfile[i].index("|")
    listSplitId=listAllfile[i][0:indexFirst]
    listSplitContent=listAllfile[i][indexFirst+1:]
    listSplitComma=listSplitContent[7:-1]#去掉SOURCE,#去掉自带的\n
    listTemp.append(listSplitId)
    listTemp.append(listSplitComma)
    listAllfileSplit.append(listTemp)
#"""
from openpyxl import *

def GetExcelRowUsed(sheet):
    rows = len(list(sheet.rows))
    print(f"=====================用了的行数rows：{rows}")
    return rows
#读取英语翻译
listAllMain=[]
wbEnglish=load_workbook("【英语】Localization-翻译.xlsx")
sheetMain=wbEnglish["Main"]
maxR=GetExcelRowUsed(sheetMain)
for i in range(2,maxR+1):#excel表格会去掉一行行头，一列列头
    tmplist = []
    for j in range(2,4):#2-3列
        str=sheetMain.cell(i,j).value
        tmplist.append(str)
    if len(tmplist)!=0:
        listAllMain.append(tmplist)
print(len(listAllMain))
wbEnglish.close()



#listAllfileSplit和listAllMain对比，找出差异文件
listDiff=[]#id,content
for aIdx in range(0,len(listAllfileSplit)):
    print(f"正在比对数据====================行号 aIdx：{aIdx}")
    listDiffTemp=[]
    for bIdx in range(0,len(listAllMain)):
        tmpId1=listAllMain[bIdx][0];
        tmpId2=listAllfileSplit[aIdx][0]
        tmpContent1=listAllMain[bIdx][1]
        tmpContent2 = listAllfileSplit[aIdx][1]
        isFind=bool(False)
        if tmpId1==tmpId2 and tmpContent1==tmpContent2:#id相等并且字符串相等
            isFind=bool(True)
            break
        elif tmpId1==tmpId2:
            isFind = bool(True)
            listDiffTemp=listAllfileSplit[aIdx]
            break

    if len(listDiffTemp)!=0 and isFind:#修改的
        listDiff.append(listDiffTemp)
    if not isFind:#新增的
        listDiffTemp=listAllfileSplit[aIdx]
        listDiff.append(listDiffTemp)
print(listDiff)

#新建差异表，用来暂存差异文件
wbDiff=Workbook()
sheetDiff=wbDiff.create_sheet("中文和ID差异表")
for idx in range(0,len(listDiff)):
    sheetDiff.cell(idx+1,1,listDiff[idx][0])
    sheetDiff.cell(idx+1,2,listDiff[idx][1])
#"""

#只存中文的，用来去重
listDiffNoRepeate=[]
sheetDiffChina=wbDiff.create_sheet("中文去重差异表")
for idx in range(0,len(listDiff)):
    if listDiff[idx][1] not in listDiffNoRepeate:
        listDiffNoRepeate.append(listDiff[idx][1])#只看中文

for idxNoRepeate in range(0,len(listDiffNoRepeate)):
    #print(f"==============================listDiffNoRepeate[idxNoRepeate]:{listDiffNoRepeate[idxNoRepeate]}")
    sheetDiffChina.cell(idxNoRepeate+1,1,listDiffNoRepeate[idxNoRepeate])
wbDiff.save("中间生成的差异表.xlsx")
wbDiff.close()

wbTranslate=load_workbook("翻译表.xlsx")
sheetTranslate=wbTranslate["中文去重差异表"]
print(f"================================向翻译文件写入不重复项==========Begin")
rows=GetExcelRowUsed(sheetTranslate)
#删除出了行头的
if rows>1:
    sheetTranslate.delete_rows(2,rows-1)
for idxNoRepeate in range(0,len(listDiffNoRepeate)):
    sheetTranslate.cell(idxNoRepeate+2,1,listDiffNoRepeate[idxNoRepeate])#从第二行开始写，第一行是行头
wbTranslate.save("翻译表.xlsx")
wbTranslate.close()
print(f"================================向翻译文件写入不重复项============End")


#cell(i,j)表示第i行，第j列的单元格数据
#更新对应的数据表格
def UpdateExcelToSave(excelBookName):
    # 读取英语翻译
    listMainInfo = []#行号，ID
    wbTemplate = load_workbook(excelBookName)
    sheetTemplateMain = wbTemplate["Main"]
    rowMax = GetExcelRowUsed(sheetTemplateMain)
    for i in range(2, rowMax + 1):  # excel表格会去掉一行行头，一列列头
        tmpTemplatelist = []
        tmpTemplatelist.append(i)  # 行号
        tmpTemplatelist.append(sheetTemplateMain.cell(i, 2).value)  # id
        if len(tmpTemplatelist) != 0:
            listMainInfo.append(tmpTemplatelist)
    lastRow=len(listMainInfo)+1#有个行头
    for idxDiff in range(0, len(listDiff)):
        isHave=bool(False)
        for idx in range(0,len(listMainInfo)):
            if listDiff[idxDiff][0]==listMainInfo[idx][1]:
                sheetTemplateMain.cell(listMainInfo[idx][0], 3).value=listDiff[idxDiff][1]#更新数据第三列
                isHave = bool(True)
                print(f"==listDiff[idxDiff][0]:{listDiff[idxDiff][0]}---listMainInfo[idx][1]:{listMainInfo[idx][1]}")
                break
        if not isHave:#插入一条数据
            lastRow=lastRow+1
            print(f"======================lastRow:{lastRow}")
            sheetTemplateMain.cell(lastRow, 1,lastRow-1)
            sheetTemplateMain.cell(lastRow, 2,listDiff[idxDiff][0])#id
            sheetTemplateMain.cell(lastRow, 3,listDiff[idxDiff][1])#中文
            sheetTemplateMain.cell(lastRow, 4,"占位")

    wbTemplate.save(excelBookName)#保存数据表格
    wbTemplate.close()
    return

# 调用三张表
print(f"===============================开始更新德语翻译文件========Begin")
UpdateExcelToSave("【德语】Localization-翻译.xlsx")
print(f"===============================开始更新德语翻译文件==========End")
print(f"===============================开始更新法语翻译文件========Begin")
UpdateExcelToSave("【法语】Localization-翻译.xlsx")
print(f"===============================开始更新法语翻译文件==========End")
print(f"===============================开始更新英语翻译文件========Begin")
UpdateExcelToSave("【英语】Localization-翻译.xlsx")
print(f"===============================开始更新英语翻译文件===========End")
#"""