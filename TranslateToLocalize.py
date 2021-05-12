#!/usr/bin/python
# -*- coding: utf-8 -*-
#第一行注释设置全局编码
import os
import builtins as built

from openpyxl import *

def GetExcelRowUsed(sheet):
    rows = len(list(sheet.rows))
    return rows
def GetTranslateByCode(codeCountry):
    listTranslate=[]
    wbTranslate = load_workbook("翻译表.xlsx")
    sheetTranslate = wbTranslate["中文去重差异表"]
    print(f"================================读取{codeCountry}翻译到列表==========Begin")
    rows=GetExcelRowUsed(sheetTranslate)
    if codeCountry=="EN":
        for idx in range(1,rows):#让出第一行行头
            listTemp1=[]
            listTemp1.append(sheetTranslate.cell(idx+1,1).value)#中文
            listTemp1.append(sheetTranslate.cell(idx+1,2).value)#英文
            listTranslate.append(listTemp1)
    elif codeCountry=="DE":
        for idx in range(1,rows):#让出第一行行头
            listTemp2=[]
            listTemp2.append(sheetTranslate.cell(idx+1,1).value)#中文
            listTemp2.append(sheetTranslate.cell(idx+1,3).value)#德文
            listTranslate.append(listTemp2)
    elif codeCountry == "FR":
        for idx in range(1,rows):#让出第一行行头
            listTemp3=[]
            listTemp3.append(sheetTranslate.cell(idx+1,1).value)#中文
            listTemp3.append(sheetTranslate.cell(idx+1,4).value)#法文
            listTranslate.append(listTemp3)
    wbTranslate.close()
    print(f"================================读取{codeCountry}翻译到列表============End")
    return listTranslate

#更新对应的数据表格
def UpdateExcelToSave(excelBookName,codeCountry):
    # 读取英语翻译
    listMainInfo = []#行号，ID
    wbTemplate = load_workbook(excelBookName)
    sheetTemplateMain = wbTemplate["Main"]
    rowMax = GetExcelRowUsed(sheetTemplateMain)
    #手机翻译表的行号和中文
    for i in range(2, rowMax + 1):  # excel表格会去掉一行行头，一列列头
        tmpTemplatelist = []
        tmpTemplatelist.append(i)  # 行号
        tmpTemplatelist.append(sheetTemplateMain.cell(i, 3).value)  #中文
        if len(tmpTemplatelist) != 0:
            listMainInfo.append(tmpTemplatelist)
    listTranslate=GetTranslateByCode(codeCountry)
    for idxTranslate in range(0, len(listTranslate)):
        for idx in range(0,len(listMainInfo)):
            if listTranslate[idxTranslate][0]==listMainInfo[idx][1]:#中文对比
                sheetTemplateMain.cell(listMainInfo[idx][0], 4).value=listTranslate[idxTranslate][1]#更新数据第四列多语言
                print(f"==更新多语言表行号:{listMainInfo[idx][0]}---内容:{listTranslate[idxTranslate][1]}")


    wbTemplate.save(excelBookName)#保存数据表格
    wbTemplate.close()
    return


def GenerateLanguage(excelBookName, codeCountry):
    # 读取英语翻译
    listMainInfo = []  # 行号，ID
    wbTemplate = load_workbook(excelBookName)
    sheetTemplateMain = wbTemplate["Main"]
    rowMax = GetExcelRowUsed(sheetTemplateMain)
    # 手机翻译表的行号和中文
    for i in range(2, rowMax + 1):  # excel表格会去掉一行行头，一列列头
        tmpTemplatelist = []
        tmpTemplatelist.append(sheetTemplateMain.cell(i, 2).value)  #id
        tmpTemplatelist.append(sheetTemplateMain.cell(i, 4).value)  #内容
        if len(tmpTemplatelist) != 0:
            listMainInfo.append(tmpTemplatelist)
    wbTemplate.close()
    #写文件

    directory=f"{codeCountry}"
    if codeCountry=="EN":
        directory="en"
    elif codeCountry=="DE":
        directory = "ge"
    elif codeCountry == "FR":
        directory = "fr"
    if not os.path.exists(directory):
        os.makedirs(directory)
    file=built.open(f"{directory}/all.txt","w",encoding="utf-8")
    for idx in range(0,len(listMainInfo)):
        str=f"{listMainInfo[idx][0]}|TRANSLATED,{listMainInfo[idx][1]}\n"
        file.write(str)
    file.close()
    return

# 调用三张表
print(f"===============================开始更新德语翻译文件========Begin")
UpdateExcelToSave("【德语】Localization-翻译.xlsx","DE")
print(f"===============================开始更新德语翻译文件==========End")
print(f"===============================开始更新法语翻译文件========Begin")
UpdateExcelToSave("【法语】Localization-翻译.xlsx","FR")
print(f"===============================开始更新法语翻译文件==========End")
print(f"===============================开始更新英语翻译文件========Begin")
UpdateExcelToSave("【英语】Localization-翻译.xlsx","EN")
print(f"===============================开始更新英语翻译文件===========End")


#生成多语言表
print(f"===============================开始写all.txt德语翻译文件========Begin")
GenerateLanguage("【德语】Localization-翻译.xlsx","DE")
print(f"===============================开始all.txt德语翻译文件==========End")
print(f"===============================开始all.txt法语翻译文件========Begin")
GenerateLanguage("【法语】Localization-翻译.xlsx","FR")
print(f"===============================开始all.txt法语翻译文件==========End")
print(f"===============================开始all.txt英语翻译文件========Begin")
GenerateLanguage("【英语】Localization-翻译.xlsx","EN")
print(f"===============================开始all.txt英语翻译文件===========End")
#"""